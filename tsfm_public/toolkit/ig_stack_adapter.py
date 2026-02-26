# Copyright contributors to the TSFM project
#
"""
IG Stack Adapter for Granite TSFM Models

Bridges the IG streaming API / DuckDB data source with all three
granite-tsfm foundation models (TinyTimeMixer, TSPulse, FlowState)
and writes results to Excel for downstream analysis alongside Chronos.

Data flow:
    DuckDB ──> DuckDBConnector ──> GraniteTSFMStack ──> ExcelResultWriter
                                       ├── TTM (point forecast)
                                       ├── FlowState (probabilistic quantiles)
                                       └── TSPulse (anomaly detection)

Install extras:
    pip install 'granite-tsfm[stack]'
"""

from __future__ import annotations

import logging
import warnings
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

import numpy as np
import pandas as pd
import torch


LOGGER = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _import_duckdb():
    """Lazy import so the module loads without duckdb installed."""
    try:
        import duckdb

        return duckdb
    except ImportError:
        raise ImportError(
            "duckdb is required for DuckDBConnector. "
            "Install with: pip install 'granite-tsfm[stack]'"
        )


def _import_openpyxl():
    """Lazy import so the module loads without openpyxl installed."""
    try:
        import openpyxl

        return openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is required for ExcelResultWriter. "
            "Install with: pip install 'granite-tsfm[stack]'"
        )


# ---------------------------------------------------------------------------
# Frequency helpers
# ---------------------------------------------------------------------------

# Maps tsfm / pandas-2.2 freq strings → old-style aliases expected by get_fixed_factor()
_TSFM_TO_FLOWSTATE: Dict[str, str] = {
    "min": "T",
    "1min": "T",
    "2min": "2T",
    "5min": "5T",
    "10min": "10T",
    "15min": "15T",
    "30min": "30T",
    "h": "H",
    "1h": "H",
    "H": "H",
    "1H": "H",
    "6h": "6H",
    "6H": "6H",
    "D": "D",
    "d": "D",
    "1D": "D",
    "1d": "D",
    "W": "W",
    "M": "M",
    "Q": "Q",
}


def _freq_to_flowstate_alias(freq: str) -> str:
    """Convert a tsfm/pandas-2.2 freq string to the old-style alias expected by
    :func:`tsfm_public.models.flowstate.utils.utils.get_fixed_factor`.

    Falls back to the original string for any unrecognised value.
    """
    return _TSFM_TO_FLOWSTATE.get(freq, freq)


# ---------------------------------------------------------------------------
# StackResult
# ---------------------------------------------------------------------------

@dataclass
class StackResult:
    """Unified output from :class:`GraniteTSFMStack`.

    Attributes:
        ttm_forecast: Point forecasts from TinyTimeMixer.
        flowstate_probabilistic: Quantile forecasts from FlowState.
        tspulse_anomaly: Anomaly scores from TSPulse.
        ensemble_summary: Blended view across all models (+ optional Chronos).
        metadata: Run metadata — model paths, timestamps, any partial errors.
    """

    ttm_forecast: Optional[pd.DataFrame]
    flowstate_probabilistic: Optional[pd.DataFrame]
    tspulse_anomaly: Optional[pd.DataFrame]
    ensemble_summary: pd.DataFrame
    metadata: Dict[str, Any] = field(default_factory=dict)


# ---------------------------------------------------------------------------
# DuckDBConnector
# ---------------------------------------------------------------------------

class DuckDBConnector:
    """Queries a DuckDB database and returns data formatted for
    :class:`~tsfm_public.toolkit.time_series_preprocessor.TimeSeriesPreprocessor`.

    The returned DataFrame always contains:
    - One datetime column (``timestamp_column``)
    - Zero or more string/int id columns (``id_columns``)
    - One or more numeric target columns (``target_columns``)

    Example::

        with DuckDBConnector(
            db_path="market.duckdb",
            table_name="ig_ticks",
            timestamp_column="timestamp",
            id_columns=["instrument_id"],
            target_columns=["mid_price"],
        ) as conn:
            df = conn.fetch_latest(context_length=512, instruments=["GBPUSD"])
    """

    def __init__(
        self,
        db_path: str,
        table_name: str,
        timestamp_column: str = "timestamp",
        id_columns: List[str] = None,
        target_columns: List[str] = None,
        extra_columns: List[str] = None,
        connection_kwargs: Dict[str, Any] = None,
    ) -> None:
        """
        Args:
            db_path: Path to the DuckDB file, or ``":memory:"`` for in-process DB.
            table_name: Name of the table (or view) to query.
            timestamp_column: Name of the datetime column.
            id_columns: Columns that distinguish separate instruments/series.
            target_columns: Numeric columns to forecast / detect anomalies on.
            extra_columns: Additional columns to pull (e.g. bid, ask spreads).
            connection_kwargs: Forwarded to ``duckdb.connect()``.
        """
        self.db_path = db_path
        self.table_name = table_name
        self.timestamp_column = timestamp_column
        self.id_columns = id_columns or ["instrument_id"]
        self.target_columns = target_columns or ["mid_price"]
        self.extra_columns = extra_columns or []
        self.connection_kwargs = connection_kwargs or {}
        self._con = None

    # ------------------------------------------------------------------
    # Connection management
    # ------------------------------------------------------------------

    def connect(self):
        """Open (and cache) a DuckDB connection."""
        if self._con is None:
            duckdb = _import_duckdb()
            self._con = duckdb.connect(self.db_path, **self.connection_kwargs)
        return self._con

    def close(self) -> None:
        """Close the DuckDB connection if open."""
        if self._con is not None:
            try:
                self._con.close()
            except Exception:
                pass
            finally:
                self._con = None

    def __enter__(self) -> "DuckDBConnector":
        self.connect()
        return self

    def __exit__(self, *args) -> None:
        self.close()

    # ------------------------------------------------------------------
    # Querying
    # ------------------------------------------------------------------

    def fetch(
        self,
        instruments: Optional[List[str]] = None,
        start_time: Optional[Union[str, pd.Timestamp]] = None,
        end_time: Optional[Union[str, pd.Timestamp]] = None,
        limit: Optional[int] = None,
        min_rows_per_instrument: int = 1,
    ) -> pd.DataFrame:
        """Execute a parameterised SELECT and return a clean DataFrame.

        Steps:
            1. Build a WHERE clause from ``instruments`` / ``start_time`` / ``end_time``.
            2. Fetch result via DuckDB's ``.df()`` method.
            3. Parse ``timestamp_column`` to ``pd.Timestamp`` if not already datetime.
            4. Sort by ``(id_columns + [timestamp_column])``.
            5. Validate that each instrument has >= ``min_rows_per_instrument`` rows.

        Args:
            instruments: Filter to specific instrument IDs (``None`` = all).
            start_time: Inclusive lower bound on timestamp.
            end_time: Inclusive upper bound on timestamp.
            limit: LIMIT clause applied **per instrument** using a window function.
            min_rows_per_instrument: Drop instruments with fewer rows (logs a warning).

        Returns:
            DataFrame with columns: ``id_columns`` + ``[timestamp_column]`` +
            ``target_columns`` + ``extra_columns``.

        Raises:
            ValueError: If the result is empty after filtering.
            RuntimeError: If the DuckDB query fails.
        """
        con = self.connect()

        all_cols = self.id_columns + [self.timestamp_column] + self.target_columns + self.extra_columns
        col_list = ", ".join(f'"{c}"' for c in all_cols)

        where_clauses = []
        if instruments:
            id_col = self.id_columns[0] if self.id_columns else None
            if id_col:
                in_list = ", ".join(f"'{i}'" for i in instruments)
                where_clauses.append(f'"{id_col}" IN ({in_list})')
        if start_time is not None:
            where_clauses.append(f'"{self.timestamp_column}" >= \'{start_time}\'')
        if end_time is not None:
            where_clauses.append(f'"{self.timestamp_column}" <= \'{end_time}\'')

        where_sql = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ""

        if limit and self.id_columns:
            id_col = self.id_columns[0]
            sql = (
                f"SELECT {col_list} FROM ("
                f"  SELECT {col_list}, "
                f"  ROW_NUMBER() OVER (PARTITION BY \"{id_col}\" ORDER BY \"{self.timestamp_column}\" DESC) AS __rn"
                f"  FROM \"{self.table_name}\" {where_sql}"
                f") WHERE __rn <= {limit}"
                f" ORDER BY {', '.join(f'\"{c}\"' for c in self.id_columns + [self.timestamp_column])}"
            )
        else:
            order_by = ", ".join(f'"{c}"' for c in self.id_columns + [self.timestamp_column])
            sql = (
                f"SELECT {col_list} FROM \"{self.table_name}\" {where_sql} ORDER BY {order_by}"
            )

        try:
            df = con.execute(sql).df()
        except Exception as exc:
            raise RuntimeError(f"DuckDB query failed: {exc}") from exc

        if df.empty:
            raise ValueError(
                f"No data returned from DuckDB query. "
                f"Table={self.table_name!r}, instruments={instruments}, "
                f"start={start_time}, end={end_time}"
            )

        # Parse timestamp column
        if not pd.api.types.is_datetime64_any_dtype(df[self.timestamp_column]):
            df[self.timestamp_column] = pd.to_datetime(df[self.timestamp_column])

        # Check for duplicate (id, timestamp) pairs
        key_cols = self.id_columns + [self.timestamp_column]
        dupes = df.duplicated(subset=key_cols)
        if dupes.any():
            raise ValueError(
                f"Duplicate (instrument, timestamp) pairs found in DuckDB result. "
                f"First duplicate: {df.loc[dupes].iloc[0].to_dict()}"
            )

        # Drop instruments with too few rows
        if self.id_columns and min_rows_per_instrument > 1:
            id_col = self.id_columns[0]
            counts = df.groupby(id_col).size()
            drop = counts[counts < min_rows_per_instrument].index.tolist()
            if drop:
                warnings.warn(
                    f"Dropping {len(drop)} instrument(s) with < {min_rows_per_instrument} rows: {drop}",
                    UserWarning,
                    stacklevel=2,
                )
                df = df[~df[id_col].isin(drop)]
            if df.empty:
                raise ValueError(
                    f"All instruments dropped due to min_rows_per_instrument={min_rows_per_instrument}."
                )

        return df.reset_index(drop=True)

    def fetch_latest(
        self,
        context_length: int,
        instruments: Optional[List[str]] = None,
    ) -> pd.DataFrame:
        """Fetch exactly the last ``context_length`` rows per instrument.

        Uses a DuckDB window function for efficiency.

        Args:
            context_length: Number of historical rows to fetch per instrument.
            instruments: Optional instrument filter.

        Returns:
            DataFrame sorted ascending by timestamp within each instrument.
        """
        return self.fetch(
            instruments=instruments,
            limit=context_length,
            min_rows_per_instrument=context_length,
        )


# ---------------------------------------------------------------------------
# GraniteTSFMStack
# ---------------------------------------------------------------------------

class GraniteTSFMStack:
    """Orchestrates TinyTimeMixer, TSPulse, and FlowState models in parallel.

    Designed to slot alongside Amazon Chronos in an ensemble: pass Chronos
    results to :meth:`run` and they will be blended into the summary sheet.

    All three model pipelines share one fitted
    :class:`~tsfm_public.toolkit.time_series_preprocessor.TimeSeriesPreprocessor`
    so that scaling is consistent across models.

    Example::

        stack = GraniteTSFMStack(
            timestamp_column="timestamp",
            id_columns=["instrument_id"],
            target_columns=["mid_price"],
            context_length=512,
            prediction_length=24,
            freq="1min",
        )
        stack.load_models()
        stack.build_preprocessor(train_df)
        result = stack.run(inference_df, chronos_results=chronos_df)
    """

    # HuggingFace Hub paths — override in __init__ with local paths if preferred
    DEFAULT_TTM_PATH = "ibm-granite/granite-timeseries-ttm-r2"
    DEFAULT_TSPULSE_PATH = "ibm-granite/granite-timeseries-tspulse-r1"
    DEFAULT_FLOWSTATE_PATH = "ibm-granite/granite-timeseries-flowstate-r1"

    def __init__(
        self,
        timestamp_column: str = "timestamp",
        id_columns: Optional[List[str]] = None,
        target_columns: Optional[List[str]] = None,
        context_length: int = 512,
        prediction_length: int = 24,
        freq: str = "1min",
        # TTM
        ttm_model_path: str = DEFAULT_TTM_PATH,
        ttm_force_return: Optional[str] = "rolling",
        ttm_add_probabilistic: bool = False,
        ttm_quantiles: Optional[List[float]] = None,
        # TSPulse
        tspulse_model_path: str = DEFAULT_TSPULSE_PATH,
        tspulse_prediction_modes: Optional[List[str]] = None,
        tspulse_aggregation_length: int = 32,
        tspulse_smoothing_length: int = 8,
        # FlowState
        flowstate_model_path: str = DEFAULT_FLOWSTATE_PATH,
        flowstate_quantiles: Optional[List[float]] = None,
        flowstate_scale_factor: Optional[float] = None,  # None = auto-compute from freq
        # Shared scaling
        scaling: bool = True,
        scaler_type: str = "standard",
        # Execution
        device: str = "cpu",
        parallel: bool = False,
        max_workers: int = 3,
        fail_on_model_error: bool = False,
    ) -> None:
        self.timestamp_column = timestamp_column
        self.id_columns = id_columns or ["instrument_id"]
        self.target_columns = target_columns or ["mid_price"]
        self.context_length = context_length
        self.prediction_length = prediction_length
        self.freq = freq

        self.ttm_model_path = ttm_model_path
        self.ttm_force_return = ttm_force_return
        self.ttm_add_probabilistic = ttm_add_probabilistic
        self.ttm_quantiles = ttm_quantiles or [0.1, 0.5, 0.9]

        self.tspulse_model_path = tspulse_model_path
        self.tspulse_prediction_modes = tspulse_prediction_modes or ["forecast"]
        self.tspulse_aggregation_length = tspulse_aggregation_length
        self.tspulse_smoothing_length = tspulse_smoothing_length

        self.flowstate_model_path = flowstate_model_path
        self.flowstate_quantiles = flowstate_quantiles

        if flowstate_scale_factor is None:
            try:
                from tsfm_public.models.flowstate.utils.utils import get_fixed_factor

                self.flowstate_scale_factor = get_fixed_factor(_freq_to_flowstate_alias(freq))
                LOGGER.info(
                    "FlowState scale_factor auto-set to %.6f for freq=%r",
                    self.flowstate_scale_factor,
                    freq,
                )
            except (NotImplementedError, Exception) as exc:
                LOGGER.warning(
                    "Could not auto-compute FlowState scale_factor for freq=%r (%s). "
                    "Defaulting to 1.0.",
                    freq,
                    exc,
                )
                self.flowstate_scale_factor = 1.0
        else:
            self.flowstate_scale_factor = max(flowstate_scale_factor, 1e-3)

        self.scaling = scaling
        self.scaler_type = scaler_type
        self.device = device
        self.parallel = parallel
        self.max_workers = max_workers
        self.fail_on_model_error = fail_on_model_error

        self._preprocessor = None
        self._ttm_model = None
        self._tspulse_model = None
        self._flowstate_model = None
        self._models_loaded = False

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def load_models(self) -> None:
        """Load all three models from HuggingFace Hub (or local paths).

        Idempotent — safe to call multiple times. Moves each model to
        ``self.device`` and sets ``eval()`` mode.
        """
        if self._models_loaded:
            return

        from tsfm_public.models.flowstate import FlowStateForPrediction
        from tsfm_public.models.tspulse import TSPulseForReconstruction
        from tsfm_public.toolkit.get_model import get_model

        # TTM — use smart revision selector
        # Normalise freq to canonical pandas alias (DEFAULT_FREQUENCY_MAPPING keys)
        # so get_model()'s frequency-token filter works correctly.
        # e.g. "1min" (not in mapping) → "min" (in mapping, R=1)
        try:
            from pandas.tseries.frequencies import to_offset

            _norm_freq = to_offset(self.freq).freqstr
        except Exception:
            _norm_freq = self.freq

        LOGGER.info("Loading TTM from %s (freq=%s) ...", self.ttm_model_path, _norm_freq)
        self._ttm_model = get_model(
            model_path=self.ttm_model_path,
            model_name="ttm",
            context_length=self.context_length,
            prediction_length=self.prediction_length,
            freq=_norm_freq,
            force_return=self.ttm_force_return,
        )
        self._ttm_model = self._ttm_model.to(self.device).eval()

        # Sync self.context_length to the actual checkpoint value.
        # get_model() selects the largest checkpoint whose context_length ≤ requested,
        # so the loaded model may have a smaller context window than self.context_length.
        # build_preprocessor() must create windows that match the checkpoint exactly.
        actual_ctx = self._ttm_model.config.context_length
        if actual_ctx != self.context_length:
            LOGGER.warning(
                "Loaded TTM checkpoint has context_length=%d but requested context_length=%d. "
                "Updating self.context_length to %d so build_preprocessor() creates "
                "windows that match the checkpoint.",
                actual_ctx,
                self.context_length,
                actual_ctx,
            )
            self.context_length = actual_ctx

        LOGGER.info(
            "TTM loaded: context=%d, prediction=%d",
            self._ttm_model.config.context_length,
            self._ttm_model.config.prediction_length,
        )

        # TSPulse — straight from_pretrained
        LOGGER.info("Loading TSPulse from %s ...", self.tspulse_model_path)
        self._tspulse_model = TSPulseForReconstruction.from_pretrained(
            self.tspulse_model_path
        )
        self._tspulse_model = self._tspulse_model.to(self.device).eval()
        LOGGER.info("TSPulse loaded.")

        # FlowState — straight from_pretrained
        LOGGER.info("Loading FlowState from %s ...", self.flowstate_model_path)
        self._flowstate_model = FlowStateForPrediction.from_pretrained(
            self.flowstate_model_path
        )
        self._flowstate_model = self._flowstate_model.to(self.device).eval()
        LOGGER.info("FlowState loaded.")

        self._models_loaded = True

    def build_preprocessor(self, data: pd.DataFrame):
        """Fit (or re-fit) the shared TimeSeriesPreprocessor on ``data``.

        Args:
            data: Raw DataFrame from :class:`DuckDBConnector`.

        Returns:
            Fitted
            :class:`~tsfm_public.toolkit.time_series_preprocessor.TimeSeriesPreprocessor`
            instance (also stored as ``self._preprocessor``).
        """
        from tsfm_public.toolkit.time_series_preprocessor import TimeSeriesPreprocessor

        tsp = TimeSeriesPreprocessor(
            id_columns=self.id_columns,
            timestamp_column=self.timestamp_column,
            target_columns=self.target_columns,
            context_length=self.context_length,
            prediction_length=self.prediction_length,
            scaling=self.scaling,
            scaler_type=self.scaler_type,
            freq=self.freq,
        )
        tsp.train(data)
        self._preprocessor = tsp
        LOGGER.info("TimeSeriesPreprocessor fitted on %d rows.", len(data))
        return tsp

    def run(
        self,
        data: pd.DataFrame,
        chronos_results: Optional[pd.DataFrame] = None,
        anomaly_threshold: float = 0.5,
    ) -> StackResult:
        """Run all three model pipelines and produce a unified :class:`StackResult`.

        Steps:
            1. Call :meth:`load_models` if models not yet loaded.
            2. Call :meth:`build_preprocessor` on the input window.
            3. Dispatch ``_run_ttm``, ``_run_tspulse``, ``_run_flowstate``
               (in parallel when ``parallel=True``).
            4. Collect outputs; on partial failure, populate
               ``metadata["model_errors"]`` or re-raise, depending on
               ``fail_on_model_error``.
            5. Build ensemble summary.

        Args:
            data: Raw market DataFrame with timestamp, id, and target columns.
            chronos_results: Optional DataFrame of Chronos point forecasts
                (columns: ``timestamp``, id columns, ``*_chronos_prediction``).
                Blended into ``ensemble_summary`` if provided.
            anomaly_threshold: TSPulse anomaly score threshold for flagging.

        Returns:
            :class:`StackResult` with four DataFrames and metadata dict.
        """
        # Validate data length
        if self.id_columns:
            id_col = self.id_columns[0]
            for inst, grp in data.groupby(id_col):
                if len(grp) < self.context_length:
                    raise ValueError(
                        f"Instrument {inst!r} has only {len(grp)} rows "
                        f"but context_length={self.context_length}. "
                        "Fetch more history or reduce context_length."
                    )
        else:
            if len(data) < self.context_length:
                raise ValueError(
                    f"Data has only {len(data)} rows "
                    f"but context_length={self.context_length}."
                )

        if not self._models_loaded:
            self.load_models()

        if self._preprocessor is None:
            self.build_preprocessor(data)

        metadata: Dict[str, Any] = {
            "run_timestamp": datetime.utcnow().isoformat(),
            "context_length": self.context_length,
            "prediction_length": self.prediction_length,
            "freq": self.freq,
            "ttm_model_path": self.ttm_model_path,
            "tspulse_model_path": self.tspulse_model_path,
            "flowstate_model_path": self.flowstate_model_path,
            "model_errors": {},
        }

        runners = {
            "ttm": self._run_ttm,
            "tspulse": self._run_tspulse,
            "flowstate": self._run_flowstate,
        }
        outputs: Dict[str, Optional[pd.DataFrame]] = {k: None for k in runners}

        if self.parallel:
            with ThreadPoolExecutor(max_workers=self.max_workers) as pool:
                futures = {pool.submit(fn, data): name for name, fn in runners.items()}
                for future in as_completed(futures):
                    name = futures[future]
                    try:
                        outputs[name] = future.result()
                    except Exception as exc:
                        if self.fail_on_model_error:
                            raise
                        metadata["model_errors"][name] = str(exc)
                        LOGGER.error("Model %s failed: %s", name, exc, exc_info=True)
        else:
            for name, fn in runners.items():
                try:
                    outputs[name] = fn(data)
                except Exception as exc:
                    if self.fail_on_model_error:
                        raise
                    metadata["model_errors"][name] = str(exc)
                    LOGGER.error("Model %s failed: %s", name, exc, exc_info=True)

        ensemble = self._build_ensemble_summary(
            ttm_df=outputs["ttm"],
            flowstate_df=outputs["flowstate"],
            tspulse_df=outputs["tspulse"],
            chronos_results=chronos_results,
            anomaly_threshold=anomaly_threshold,
        )

        return StackResult(
            ttm_forecast=outputs["ttm"],
            flowstate_probabilistic=outputs["flowstate"],
            tspulse_anomaly=outputs["tspulse"],
            ensemble_summary=ensemble,
            metadata=metadata,
        )

    # ------------------------------------------------------------------
    # Internal pipeline runners
    # ------------------------------------------------------------------

    def _run_ttm(self, data: pd.DataFrame) -> pd.DataFrame:
        """Run TinyTimeMixer forecasting pipeline."""
        from tsfm_public.toolkit.recursive_predictor import RecursivePredictor, RecursivePredictorConfig
        from tsfm_public.toolkit.time_series_forecasting_pipeline import TimeSeriesForecastingPipeline

        prob_processor = None
        if self.ttm_add_probabilistic:
            try:
                from tsfm_public.toolkit.conformal import PostHocProbabilisticProcessor

                prob_processor = PostHocProbabilisticProcessor(quantiles=self.ttm_quantiles)
            except Exception as exc:
                LOGGER.warning("Could not build probabilistic processor: %s", exc)

        # Wrap in RecursivePredictor if the loaded model's prediction_length < requested
        ttm_pred_len = getattr(self._ttm_model.config, "prediction_length", self.prediction_length)
        if ttm_pred_len < self.prediction_length:
            LOGGER.info(
                "TTM prediction_length=%d < requested=%d; wrapping with RecursivePredictor.",
                ttm_pred_len,
                self.prediction_length,
            )
            rp_config = RecursivePredictorConfig(
                model=self._ttm_model,
                requested_prediction_length=self.prediction_length,
                model_prediction_length=ttm_pred_len,
                loss="mse",
            )
            effective_model = RecursivePredictor(rp_config)
        else:
            effective_model = self._ttm_model

        pipeline = TimeSeriesForecastingPipeline(
            model=effective_model,
            feature_extractor=self._preprocessor,
            probabilistic_processor=prob_processor,
            device=self.device,
        )
        return pipeline(data, batch_size=64, num_workers=0)

    def _run_tspulse(self, data: pd.DataFrame) -> pd.DataFrame:
        """Run TSPulse anomaly detection pipeline."""
        from tsfm_public.toolkit.time_series_anomaly_detection_pipeline import (
            TimeSeriesAnomalyDetectionPipeline,
        )

        pipeline = TimeSeriesAnomalyDetectionPipeline(
            model=self._tspulse_model,
            feature_extractor=self._preprocessor,
            prediction_mode=self.tspulse_prediction_modes,
            aggregation_length=self.tspulse_aggregation_length,
            smoothing_length=self.tspulse_smoothing_length,
            device=self.device,
        )
        return pipeline(data, batch_size=64, num_workers=0)

    def _run_flowstate(self, data: pd.DataFrame) -> pd.DataFrame:
        """Run FlowState probabilistic forecasting.

        FlowState is **univariate-only**. We loop over each instrument and
        each target column, running one forward pass per combination and
        assembling a tidy quantile-forecast DataFrame.
        """
        result_rows = []

        # Group by instrument if id_columns present
        if self.id_columns:
            groups = data.groupby(self.id_columns[0])
        else:
            groups = [("_single_", data)]

        for inst_id, grp in groups:
            grp = grp.sort_values(self.timestamp_column)

            for target_col in self.target_columns:
                series = grp[target_col].values.astype(np.float32)

                # Take last context_length rows
                if len(series) >= self.context_length:
                    series = series[-self.context_length:]
                else:
                    # Zero-pad on the left
                    pad = np.zeros(self.context_length - len(series), dtype=np.float32)
                    series = np.concatenate([pad, series])

                # Shape: (1, context_length, 1) with batch_first=True
                tensor = torch.tensor(series, dtype=torch.float32).unsqueeze(0).unsqueeze(-1)
                tensor = tensor.to(self.device)

                with torch.no_grad():
                    out = self._flowstate_model(
                        past_values=tensor,
                        prediction_length=self.prediction_length,
                        scale_factor=self.flowstate_scale_factor,
                        prediction_type="quantile",
                        batch_first=True,
                    )

                # prediction_outputs: (n_preds, batch, n_quantiles, pred_len, n_ch)
                # After _combine_cpm_predictions in forward, shape is (batch, n_quantiles, pred_len)
                pred = out.prediction_outputs  # (batch, n_quantiles, pred_len)
                if pred.dim() == 5:
                    # (n_preds, batch, n_quantiles, pred_len, n_ch) -> squeeze
                    pred = pred[0, :, :, :, 0]  # (batch, n_quantiles, pred_len)
                pred = pred.squeeze(0).cpu().numpy()  # (n_quantiles, pred_len)

                # Build future timestamps from last observed timestamp
                last_ts = grp[self.timestamp_column].iloc[-1]
                freq_offset = pd.tseries.frequencies.to_offset(self.freq)
                future_ts = pd.date_range(
                    start=last_ts + freq_offset,
                    periods=self.prediction_length,
                    freq=self.freq,
                )

                # Determine quantile labels
                if self.flowstate_quantiles is not None:
                    quantile_labels = [f"q{int(q * 100):02d}" for q in self.flowstate_quantiles]
                    n_q = min(len(quantile_labels), pred.shape[0])
                    pred = pred[:n_q]
                    quantile_labels = quantile_labels[:n_q]
                else:
                    n_q = pred.shape[0]
                    quantile_labels = [f"q{int(100 * (i + 1) / (n_q + 1)):02d}" for i in range(n_q)]

                for step in range(self.prediction_length):
                    row: Dict[str, Any] = {self.timestamp_column: future_ts[step]}
                    if self.id_columns:
                        row[self.id_columns[0]] = inst_id
                    for qi, ql in enumerate(quantile_labels):
                        row[f"{target_col}_{ql}"] = float(pred[qi, step])
                    result_rows.append(row)

        if not result_rows:
            return pd.DataFrame()

        return pd.DataFrame(result_rows)

    def _build_ensemble_summary(
        self,
        ttm_df: Optional[pd.DataFrame],
        flowstate_df: Optional[pd.DataFrame],
        tspulse_df: Optional[pd.DataFrame],
        chronos_results: Optional[pd.DataFrame],
        anomaly_threshold: float,
    ) -> pd.DataFrame:
        """Blend model outputs into a single ensemble summary DataFrame.

        Columns in the result:
        - ``ensemble_point``    — mean of available point forecasts
        - ``ensemble_lower``    — minimum lower quantile across models
        - ``ensemble_upper``    — maximum upper quantile across models
        - ``model_agreement``   — std-dev of point forecasts (divergence signal)
        - ``models_contributing`` — count of models that produced a forecast
        - ``anomaly_flagged``   — True where TSPulse anomaly_score >= threshold
        """
        join_keys = self.id_columns + [self.timestamp_column]

        frames = []

        if ttm_df is not None and not ttm_df.empty:
            pred_cols = [c for c in ttm_df.columns if c.endswith("_prediction")]
            if pred_cols:
                sub = ttm_df[join_keys + pred_cols].copy()
                sub.rename(
                    columns={c: f"ttm_{c}" for c in pred_cols},
                    inplace=True,
                )
                frames.append(sub)

        if flowstate_df is not None and not flowstate_df.empty:
            q_cols = [c for c in flowstate_df.columns if "_q" in c]
            if q_cols:
                sub = flowstate_df[join_keys + q_cols].copy()
                sub.rename(
                    columns={c: f"flowstate_{c}" for c in q_cols},
                    inplace=True,
                )
                frames.append(sub)

        if chronos_results is not None and not chronos_results.empty:
            chron_pred_cols = [
                c for c in chronos_results.columns if c not in join_keys
            ]
            sub = chronos_results[join_keys + chron_pred_cols].copy()
            frames.append(sub)

        if not frames:
            return pd.DataFrame(columns=join_keys + ["ensemble_point", "models_contributing"])

        ensemble = frames[0]
        for f in frames[1:]:
            ensemble = ensemble.merge(f, on=join_keys, how="outer")

        # Identify point forecast columns for each model
        point_forecast_cols = []
        for col in ensemble.columns:
            if col in join_keys:
                continue
            if "_prediction" in col or "_chronos" in col:
                point_forecast_cols.append(col)
            # FlowState q50 as point estimate
            if "_q50" in col and col.startswith("flowstate_"):
                point_forecast_cols.append(col)

        point_forecast_cols = list(dict.fromkeys(point_forecast_cols))  # deduplicate

        if point_forecast_cols:
            ensemble["ensemble_point"] = ensemble[point_forecast_cols].mean(axis=1)
            ensemble["model_agreement"] = ensemble[point_forecast_cols].std(axis=1).fillna(0.0)
            ensemble["models_contributing"] = ensemble[point_forecast_cols].notna().sum(axis=1)
        else:
            ensemble["ensemble_point"] = np.nan
            ensemble["model_agreement"] = np.nan
            ensemble["models_contributing"] = 0

        # Uncertainty envelope
        lower_cols = [c for c in ensemble.columns if "_q10" in c]
        upper_cols = [c for c in ensemble.columns if "_q90" in c]
        ensemble["ensemble_lower"] = ensemble[lower_cols].min(axis=1) if lower_cols else np.nan
        ensemble["ensemble_upper"] = ensemble[upper_cols].max(axis=1) if upper_cols else np.nan

        # Merge anomaly scores
        if tspulse_df is not None and not tspulse_df.empty and "anomaly_score" in tspulse_df.columns:
            anon_cols = join_keys + ["anomaly_score"]
            anomaly_sub = tspulse_df[anon_cols].drop_duplicates()
            ensemble = ensemble.merge(anomaly_sub, on=join_keys, how="left")
            ensemble["anomaly_flagged"] = ensemble["anomaly_score"] >= anomaly_threshold
        else:
            ensemble["anomaly_flagged"] = False

        return ensemble.reset_index(drop=True)


# ---------------------------------------------------------------------------
# ExcelResultWriter
# ---------------------------------------------------------------------------

class ExcelResultWriter:
    """Writes :class:`StackResult` DataFrames to a multi-sheet Excel workbook.

    Sheet layout:
    - ``TTM_Forecast``          — TinyTimeMixer point (+ optional quantile) forecasts
    - ``FlowState_Probabilistic`` — FlowState quantile forecast columns
    - ``TSPulse_Anomaly``       — Anomaly scores per timestamp/instrument
    - ``Ensemble_Summary``      — Blended view across all models

    When ``overwrite=False`` (default), new rows are **appended** below
    existing data in each sheet, enabling multi-run history in one file.

    Example::

        writer = ExcelResultWriter("ig_results.xlsx")
        writer.write(result, run_label="2025-01-15T09:00Z")
    """

    SHEET_NAMES = {
        "ttm_forecast": "TTM_Forecast",
        "flowstate_probabilistic": "FlowState_Probabilistic",
        "tspulse_anomaly": "TSPulse_Anomaly",
        "ensemble_summary": "Ensemble_Summary",
    }

    def __init__(
        self,
        output_path: Union[str, Path],
        overwrite: bool = False,
        freeze_panes: bool = True,
        auto_filter: bool = True,
    ) -> None:
        """
        Args:
            output_path: Path to the ``.xlsx`` file.
            overwrite: If ``True`` overwrite file on each write; if ``False``
                append new rows below existing data.
            freeze_panes: Freeze the header row in each sheet.
            auto_filter: Apply Excel AutoFilter to each sheet.
        """
        self.output_path = Path(output_path)
        self.overwrite = overwrite
        self.freeze_panes = freeze_panes
        self.auto_filter = auto_filter

    def write(
        self,
        result: StackResult,
        run_label: Optional[str] = None,
    ) -> Path:
        """Write all four sheets from a :class:`StackResult` to the Excel file.

        Args:
            result: :class:`StackResult` from :meth:`GraniteTSFMStack.run`.
            run_label: Optional string label prepended as a column to each
                sheet (defaults to current UTC ISO timestamp).

        Returns:
            Resolved path to the written file.
        """
        openpyxl = _import_openpyxl()

        label = run_label or datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")

        sheet_data = {
            "TTM_Forecast": result.ttm_forecast,
            "FlowState_Probabilistic": result.flowstate_probabilistic,
            "TSPulse_Anomaly": result.tspulse_anomaly,
            "Ensemble_Summary": result.ensemble_summary,
        }

        file_exists = self.output_path.exists()

        if not file_exists or self.overwrite:
            wb = openpyxl.Workbook()
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
            existing_data: Dict[str, List] = {name: [] for name in sheet_data}
        else:
            wb = openpyxl.load_workbook(str(self.output_path))
            existing_data = {}
            for sheet_name in sheet_data:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows = list(ws.values)
                    existing_data[sheet_name] = rows
                else:
                    existing_data[sheet_name] = []

        for sheet_name, df in sheet_data.items():
            self._write_sheet(wb, sheet_name, df, label, existing_data.get(sheet_name, []))

        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(self.output_path))
        LOGGER.info("Results written to %s", self.output_path.resolve())
        return self.output_path.resolve()

    def _write_sheet(
        self,
        wb,
        sheet_name: str,
        df: Optional[pd.DataFrame],
        run_label: str,
        existing_rows: List,
    ) -> None:
        """Write (or replace) a single named sheet with ``run_label`` column."""
        openpyxl = _import_openpyxl()

        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(title=sheet_name)

        # Build rows to write
        if df is None or (isinstance(df, pd.DataFrame) and df.empty):
            # Placeholder when model failed or produced no output
            ws.append(["run_label", "status"])
            ws.append([run_label, "Model unavailable — see StackResult.metadata for error details."])
        else:
            new_df = df.copy()
            new_df.insert(0, "run_label", run_label)

            # Write existing rows first (skip header if present)
            header_written = False
            if existing_rows:
                for i, row in enumerate(existing_rows):
                    if i == 0:
                        # Use the header from existing data
                        ws.append(list(row))
                        header_written = True
                    else:
                        ws.append(list(row))

            # Write header if not already written
            if not header_written:
                ws.append(list(new_df.columns))

            # Write new data rows
            for row_tuple in new_df.itertuples(index=False, name=None):
                ws.append(list(row_tuple))

        # Apply formatting
        if ws.max_row > 1:
            if self.freeze_panes:
                ws.freeze_panes = "A2"
            if self.auto_filter:
                ws.auto_filter.ref = ws.dimensions


# ---------------------------------------------------------------------------
# Public module exports
# ---------------------------------------------------------------------------

__all__ = [
    "DuckDBConnector",
    "GraniteTSFMStack",
    "ExcelResultWriter",
    "StackResult",
]

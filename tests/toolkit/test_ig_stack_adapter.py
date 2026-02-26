# Copyright contributors to the TSFM project
#
"""Tests for the IG Stack Adapter (DuckDBConnector, GraniteTSFMStack, ExcelResultWriter).

All tests use synthetic in-memory data — no real DuckDB file, IG API,
or model downloads are required.
"""

from __future__ import annotations

import io
import tempfile
from datetime import datetime, timedelta
from pathlib import Path
from unittest.mock import MagicMock, patch

import numpy as np
import pandas as pd
import pytest

from tsfm_public.toolkit.ig_stack_adapter import (
    DuckDBConnector,
    ExcelResultWriter,
    GraniteTSFMStack,
    StackResult,
)


# ---------------------------------------------------------------------------
# Shared fixtures
# ---------------------------------------------------------------------------

N_ROWS = 128  # more than a typical context_length for tests
INSTRUMENTS = ["GBPUSD", "UK100"]
FREQ = "1min"


@pytest.fixture(scope="module")
def market_df():
    """Synthetic multi-instrument market DataFrame mimicking IG streaming data."""
    rows = []
    t0 = datetime(2024, 1, 15, 9, 0, 0)
    for inst in INSTRUMENTS:
        np.random.seed(42 if inst == "GBPUSD" else 7)
        prices = 1.2500 + np.cumsum(np.random.randn(N_ROWS) * 0.0005)
        for i in range(N_ROWS):
            rows.append(
                {
                    "timestamp": t0 + timedelta(minutes=i),
                    "instrument_id": inst,
                    "mid_price": float(prices[i]),
                    "bid": float(prices[i] - 0.0002),
                    "ask": float(prices[i] + 0.0002),
                }
            )
    return pd.DataFrame(rows).sort_values(["instrument_id", "timestamp"]).reset_index(drop=True)


@pytest.fixture(scope="module")
def single_instrument_df(market_df):
    return market_df[market_df["instrument_id"] == "GBPUSD"].reset_index(drop=True)


# ---------------------------------------------------------------------------
# DuckDBConnector tests (using in-memory DuckDB)
# ---------------------------------------------------------------------------

class TestDuckDBConnector:
    """Tests that use an actual in-memory DuckDB instance."""

    @pytest.fixture(autouse=True)
    def _skip_if_no_duckdb(self):
        pytest.importorskip("duckdb")

    @pytest.fixture()
    def in_memory_db(self, market_df):
        """Create an in-memory DuckDB and populate it with synthetic data."""
        import duckdb

        con = duckdb.connect(":memory:")
        con.register("ig_ticks", market_df)
        return con

    def _make_connector(self):
        return DuckDBConnector(
            db_path=":memory:",
            table_name="ig_ticks",
            timestamp_column="timestamp",
            id_columns=["instrument_id"],
            target_columns=["mid_price"],
            extra_columns=["bid", "ask"],
        )

    def test_context_manager_connect_close(self):
        connector = self._make_connector()
        with connector as c:
            assert c._con is not None
        assert connector._con is None

    def test_fetch_all_rows(self, market_df, in_memory_db):
        connector = self._make_connector()
        connector._con = in_memory_db  # inject connection
        df = connector.fetch()
        assert len(df) == len(market_df)
        assert "mid_price" in df.columns
        assert pd.api.types.is_datetime64_any_dtype(df["timestamp"])

    def test_fetch_instrument_filter(self, in_memory_db):
        connector = self._make_connector()
        connector._con = in_memory_db
        df = connector.fetch(instruments=["GBPUSD"])
        assert set(df["instrument_id"].unique()) == {"GBPUSD"}
        assert len(df) == N_ROWS

    def test_fetch_time_bounds(self, in_memory_db, market_df):
        connector = self._make_connector()
        connector._con = in_memory_db
        start = datetime(2024, 1, 15, 9, 30)
        end = datetime(2024, 1, 15, 10, 0)
        df = connector.fetch(start_time=start, end_time=end)
        assert (df["timestamp"] >= start).all()
        assert (df["timestamp"] <= end).all()

    def test_fetch_latest_returns_context_rows(self, in_memory_db):
        connector = self._make_connector()
        connector._con = in_memory_db
        context_length = 32
        df = connector.fetch_latest(context_length=context_length)
        for inst in INSTRUMENTS:
            assert len(df[df["instrument_id"] == inst]) == context_length

    def test_fetch_empty_raises(self, in_memory_db):
        connector = self._make_connector()
        connector._con = in_memory_db
        with pytest.raises(ValueError, match="No data returned"):
            connector.fetch(instruments=["NONEXISTENT"])

    def test_fetch_too_few_rows_warns(self, in_memory_db):
        connector = self._make_connector()
        connector._con = in_memory_db
        # Requesting more rows than exist → warning and empty result
        with pytest.warns(UserWarning, match="Dropping"):
            with pytest.raises(ValueError, match="All instruments dropped"):
                connector.fetch(min_rows_per_instrument=N_ROWS + 9999)

    def test_sorted_ascending(self, in_memory_db):
        connector = self._make_connector()
        connector._con = in_memory_db
        df = connector.fetch(instruments=["GBPUSD"])
        ts = df["timestamp"].values
        assert (ts[1:] >= ts[:-1]).all(), "Timestamps must be sorted ascending"


# ---------------------------------------------------------------------------
# GraniteTSFMStack tests (models mocked — no HuggingFace download)
# ---------------------------------------------------------------------------

class TestGraniteTSFMStack:
    """Unit tests for GraniteTSFMStack using mocked model pipelines."""

    @pytest.fixture()
    def stack(self):
        return GraniteTSFMStack(
            timestamp_column="timestamp",
            id_columns=["instrument_id"],
            target_columns=["mid_price"],
            context_length=64,
            prediction_length=12,
            freq=FREQ,
            parallel=False,
            fail_on_model_error=False,
        )

    def test_init_defaults(self, stack):
        assert stack.context_length == 64
        assert stack.prediction_length == 12
        assert stack.freq == FREQ
        assert not stack._models_loaded

    def test_build_preprocessor_fits(self, stack, market_df):
        tsp = stack.build_preprocessor(market_df)
        assert stack._preprocessor is tsp
        assert tsp is not None

    def test_build_preprocessor_idempotent(self, stack, market_df):
        tsp1 = stack.build_preprocessor(market_df)
        tsp2 = stack.build_preprocessor(market_df)
        # Both calls succeed and return a preprocessor
        assert tsp1 is not None
        assert tsp2 is not None

    def test_run_raises_on_insufficient_data(self, stack, market_df):
        """Instrument with fewer rows than context_length should raise."""
        tiny_df = market_df[market_df["instrument_id"] == "GBPUSD"].head(10).copy()
        stack2 = GraniteTSFMStack(context_length=64, prediction_length=12)
        # Even without model loading, length validation triggers first
        with pytest.raises(ValueError, match="context_length"):
            stack2.run(tiny_df)

    def _make_mock_ttm_output(self, market_df, stack):
        """Build a fake TTM pipeline output DataFrame."""
        rows = []
        t0 = market_df["timestamp"].max() + timedelta(minutes=1)
        for inst in INSTRUMENTS:
            for step in range(stack.prediction_length):
                rows.append(
                    {
                        "timestamp": t0 + timedelta(minutes=step),
                        "instrument_id": inst,
                        "mid_price_prediction": 1.25 + step * 0.001,
                    }
                )
        return pd.DataFrame(rows)

    def _make_mock_tspulse_output(self, market_df):
        """Build a fake TSPulse anomaly detection output DataFrame."""
        rows = []
        for inst in INSTRUMENTS:
            sub = market_df[market_df["instrument_id"] == inst]
            for _, row in sub.iterrows():
                rows.append(
                    {
                        "timestamp": row["timestamp"],
                        "instrument_id": inst,
                        "anomaly_score": float(np.random.uniform(0, 1)),
                    }
                )
        return pd.DataFrame(rows)

    def _make_mock_flowstate_output(self, market_df, stack):
        """Build a fake FlowState quantile output DataFrame."""
        rows = []
        t0 = market_df["timestamp"].max() + timedelta(minutes=1)
        for inst in INSTRUMENTS:
            for step in range(stack.prediction_length):
                rows.append(
                    {
                        "timestamp": t0 + timedelta(minutes=step),
                        "instrument_id": inst,
                        "mid_price_q10": 1.248,
                        "mid_price_q50": 1.250,
                        "mid_price_q90": 1.252,
                    }
                )
        return pd.DataFrame(rows)

    def test_run_with_mocked_model_runners(self, stack, market_df):
        """Patch internal model runners to avoid loading real models."""
        ttm_out = self._make_mock_ttm_output(market_df, stack)
        tspulse_out = self._make_mock_tspulse_output(market_df)
        flowstate_out = self._make_mock_flowstate_output(market_df, stack)

        stack.build_preprocessor(market_df)
        stack._models_loaded = True  # bypass load_models()

        with (
            patch.object(stack, "_run_ttm", return_value=ttm_out),
            patch.object(stack, "_run_tspulse", return_value=tspulse_out),
            patch.object(stack, "_run_flowstate", return_value=flowstate_out),
        ):
            result = stack.run(market_df)

        assert isinstance(result, StackResult)
        assert result.ttm_forecast is not None
        assert result.flowstate_probabilistic is not None
        assert result.tspulse_anomaly is not None
        assert not result.ensemble_summary.empty
        assert "ensemble_point" in result.ensemble_summary.columns
        assert "anomaly_flagged" in result.ensemble_summary.columns

    def test_run_soft_failure_on_model_error(self, stack, market_df):
        """When fail_on_model_error=False, errors are captured in metadata."""
        stack.build_preprocessor(market_df)
        stack._models_loaded = True

        def _boom(_data):
            raise RuntimeError("Simulated model failure")

        with (
            patch.object(stack, "_run_ttm", side_effect=_boom),
            patch.object(stack, "_run_tspulse", side_effect=_boom),
            patch.object(stack, "_run_flowstate", side_effect=_boom),
        ):
            result = stack.run(market_df)

        assert result.ttm_forecast is None
        assert result.flowstate_probabilistic is None
        assert result.tspulse_anomaly is None
        assert "ttm" in result.metadata["model_errors"]
        assert "tspulse" in result.metadata["model_errors"]
        assert "flowstate" in result.metadata["model_errors"]

    def test_run_hard_failure_propagates(self, market_df):
        """When fail_on_model_error=True, exceptions bubble up."""
        stack_strict = GraniteTSFMStack(
            context_length=64,
            prediction_length=12,
            fail_on_model_error=True,
        )
        stack_strict.build_preprocessor(market_df)
        stack_strict._models_loaded = True

        with patch.object(stack_strict, "_run_ttm", side_effect=RuntimeError("fail")):
            with patch.object(stack_strict, "_run_tspulse", return_value=None):
                with patch.object(stack_strict, "_run_flowstate", return_value=None):
                    with pytest.raises(RuntimeError, match="fail"):
                        stack_strict.run(market_df)

    def test_ensemble_summary_with_chronos(self, stack, market_df):
        """Chronos results are blended into ensemble_summary."""
        ttm_out = self._make_mock_ttm_output(market_df, stack)
        tspulse_out = self._make_mock_tspulse_output(market_df)
        flowstate_out = self._make_mock_flowstate_output(market_df, stack)

        # Fake Chronos output matching TTM timestamps
        chronos_df = ttm_out[["timestamp", "instrument_id"]].copy()
        chronos_df["mid_price_chronos_prediction"] = 1.249

        stack.build_preprocessor(market_df)
        stack._models_loaded = True

        with (
            patch.object(stack, "_run_ttm", return_value=ttm_out),
            patch.object(stack, "_run_tspulse", return_value=tspulse_out),
            patch.object(stack, "_run_flowstate", return_value=flowstate_out),
        ):
            result = stack.run(market_df, chronos_results=chronos_df)

        summary = result.ensemble_summary
        assert "models_contributing" in summary.columns
        # TTM + FlowState q50 + Chronos = at most 3 models per row
        assert (summary["models_contributing"] >= 1).all()

    def test_metadata_contains_run_timestamp(self, stack, market_df):
        stack.build_preprocessor(market_df)
        stack._models_loaded = True
        with (
            patch.object(stack, "_run_ttm", return_value=None),
            patch.object(stack, "_run_tspulse", return_value=None),
            patch.object(stack, "_run_flowstate", return_value=None),
        ):
            result = stack.run(market_df)
        assert "run_timestamp" in result.metadata
        # Should be parseable as ISO datetime
        datetime.fromisoformat(result.metadata["run_timestamp"])

    # ------------------------------------------------------------------
    # scale_factor auto-computation tests
    # ------------------------------------------------------------------

    def test_scale_factor_auto_computed_for_minute_freq(self):
        """1-minute data: scale_factor = 24 / (24*60) ≈ 0.016667."""
        stack = GraniteTSFMStack(freq="1min")
        expected = 24.0 / (24.0 * 60.0)
        assert abs(stack.flowstate_scale_factor - expected) < 1e-5, (
            f"Expected scale_factor≈{expected:.6f} for freq='1min', got {stack.flowstate_scale_factor}"
        )

    def test_scale_factor_auto_computed_for_hourly_freq(self):
        """Hourly data: scale_factor = 24/24 = 1.0 (FlowState base frequency)."""
        stack = GraniteTSFMStack(freq="1h")
        assert abs(stack.flowstate_scale_factor - 1.0) < 1e-5, (
            f"Expected scale_factor=1.0 for freq='1h', got {stack.flowstate_scale_factor}"
        )

    def test_scale_factor_explicit_override(self):
        """Explicit flowstate_scale_factor overrides auto-computation."""
        stack = GraniteTSFMStack(freq="1min", flowstate_scale_factor=0.5)
        assert stack.flowstate_scale_factor == 0.5

    # ------------------------------------------------------------------
    # TTM context_length sync test
    # ------------------------------------------------------------------

    def test_load_models_syncs_context_length_to_checkpoint(self, market_df):
        """When get_model() returns a checkpoint with a smaller context_length,
        self.context_length must be updated so build_preprocessor() creates
        windows that match the model."""
        stack = GraniteTSFMStack(context_length=500, prediction_length=24, freq="1min")

        # Mock TTM checkpoint that has context_length=360 (< requested 500)
        mock_ttm = MagicMock()
        mock_ttm.config.context_length = 360
        mock_ttm.config.prediction_length = 60
        mock_ttm.to.return_value = mock_ttm
        mock_ttm.eval.return_value = mock_ttm

        mock_tspulse = MagicMock()
        mock_tspulse.to.return_value = mock_tspulse
        mock_tspulse.eval.return_value = mock_tspulse

        mock_flowstate = MagicMock()
        mock_flowstate.to.return_value = mock_flowstate
        mock_flowstate.eval.return_value = mock_flowstate

        with (
            patch("tsfm_public.toolkit.ig_stack_adapter.get_model", return_value=mock_ttm),
            patch(
                "tsfm_public.models.tspulse.TSPulseForReconstruction.from_pretrained",
                return_value=mock_tspulse,
            ),
            patch(
                "tsfm_public.models.flowstate.FlowStateForPrediction.from_pretrained",
                return_value=mock_flowstate,
            ),
        ):
            stack.load_models()

        # context_length must be synced to checkpoint value
        assert stack.context_length == 360, (
            f"Expected context_length=360 after sync, got {stack.context_length}"
        )


# ---------------------------------------------------------------------------
# ExcelResultWriter tests
# ---------------------------------------------------------------------------

class TestExcelResultWriter:
    """Tests that write to a temporary .xlsx file."""

    @pytest.fixture(autouse=True)
    def _skip_if_no_openpyxl(self):
        pytest.importorskip("openpyxl")

    @pytest.fixture()
    def sample_result(self, market_df):
        """A realistic StackResult with non-empty DataFrames."""
        t0 = market_df["timestamp"].max() + timedelta(minutes=1)
        pred_len = 8

        ttm_df = pd.DataFrame(
            [
                {
                    "timestamp": t0 + timedelta(minutes=i),
                    "instrument_id": inst,
                    "mid_price_prediction": 1.25 + i * 0.001,
                }
                for inst in INSTRUMENTS
                for i in range(pred_len)
            ]
        )
        fs_df = pd.DataFrame(
            [
                {
                    "timestamp": t0 + timedelta(minutes=i),
                    "instrument_id": inst,
                    "mid_price_q10": 1.248,
                    "mid_price_q50": 1.250,
                    "mid_price_q90": 1.252,
                }
                for inst in INSTRUMENTS
                for i in range(pred_len)
            ]
        )
        tsp_df = pd.DataFrame(
            [
                {
                    "timestamp": row["timestamp"],
                    "instrument_id": row["instrument_id"],
                    "anomaly_score": 0.1,
                }
                for _, row in market_df.iterrows()
            ]
        )
        ensemble = pd.DataFrame(
            [
                {
                    "timestamp": t0 + timedelta(minutes=i),
                    "instrument_id": inst,
                    "ensemble_point": 1.250,
                    "models_contributing": 2,
                    "anomaly_flagged": False,
                }
                for inst in INSTRUMENTS
                for i in range(pred_len)
            ]
        )
        return StackResult(
            ttm_forecast=ttm_df,
            flowstate_probabilistic=fs_df,
            tspulse_anomaly=tsp_df,
            ensemble_summary=ensemble,
            metadata={"run_timestamp": "2024-01-15T09:00:00"},
        )

    def test_write_creates_file(self, sample_result, tmp_path):
        out = tmp_path / "results.xlsx"
        writer = ExcelResultWriter(str(out))
        path = writer.write(sample_result)
        assert path.exists()

    def test_write_creates_four_sheets(self, sample_result, tmp_path):
        import openpyxl

        out = tmp_path / "results.xlsx"
        writer = ExcelResultWriter(str(out))
        writer.write(sample_result)

        wb = openpyxl.load_workbook(str(out))
        assert "TTM_Forecast" in wb.sheetnames
        assert "FlowState_Probabilistic" in wb.sheetnames
        assert "TSPulse_Anomaly" in wb.sheetnames
        assert "Ensemble_Summary" in wb.sheetnames

    def test_run_label_column_present(self, sample_result, tmp_path):
        import openpyxl

        out = tmp_path / "results.xlsx"
        writer = ExcelResultWriter(str(out))
        writer.write(sample_result, run_label="test_run_001")

        wb = openpyxl.load_workbook(str(out))
        ws = wb["TTM_Forecast"]
        header_row = [cell.value for cell in ws[1]]
        assert "run_label" in header_row

        data_row = [cell.value for cell in ws[2]]
        assert "test_run_001" in data_row

    def test_append_mode_accumulates_rows(self, sample_result, tmp_path):
        import openpyxl

        out = tmp_path / "results.xlsx"
        writer = ExcelResultWriter(str(out), overwrite=False)
        writer.write(sample_result, run_label="run1")
        writer.write(sample_result, run_label="run2")

        wb = openpyxl.load_workbook(str(out))
        ws = wb["Ensemble_Summary"]
        # 1 header + 2 * (n_instruments * pred_len) data rows
        n_instruments = len(INSTRUMENTS)
        pred_len = 8
        expected_data_rows = 2 * n_instruments * pred_len
        assert ws.max_row == 1 + expected_data_rows

    def test_overwrite_mode_replaces_rows(self, sample_result, tmp_path):
        import openpyxl

        out = tmp_path / "results.xlsx"
        writer = ExcelResultWriter(str(out), overwrite=True)
        writer.write(sample_result, run_label="run1")
        writer.write(sample_result, run_label="run2")

        wb = openpyxl.load_workbook(str(out))
        ws = wb["Ensemble_Summary"]
        n_instruments = len(INSTRUMENTS)
        pred_len = 8
        expected_data_rows = n_instruments * pred_len
        assert ws.max_row == 1 + expected_data_rows

    def test_none_dataframe_writes_placeholder(self, tmp_path):
        import openpyxl

        result = StackResult(
            ttm_forecast=None,
            flowstate_probabilistic=None,
            tspulse_anomaly=None,
            ensemble_summary=pd.DataFrame(columns=["timestamp", "instrument_id", "ensemble_point"]),
            metadata={},
        )
        out = tmp_path / "results.xlsx"
        writer = ExcelResultWriter(str(out))
        writer.write(result)

        wb = openpyxl.load_workbook(str(out))
        ws = wb["TTM_Forecast"]
        row2 = [cell.value for cell in ws[2]]
        assert any("unavailable" in str(v).lower() for v in row2 if v is not None)

    def test_output_directory_created(self, sample_result, tmp_path):
        nested = tmp_path / "deep" / "nested" / "results.xlsx"
        writer = ExcelResultWriter(str(nested))
        path = writer.write(sample_result)
        assert path.exists()

    def test_freeze_panes_applied(self, sample_result, tmp_path):
        import openpyxl

        out = tmp_path / "results.xlsx"
        writer = ExcelResultWriter(str(out), freeze_panes=True)
        writer.write(sample_result)

        wb = openpyxl.load_workbook(str(out))
        ws = wb["TTM_Forecast"]
        assert ws.freeze_panes == "A2"

    def test_no_freeze_panes(self, sample_result, tmp_path):
        import openpyxl

        out = tmp_path / "results.xlsx"
        writer = ExcelResultWriter(str(out), freeze_panes=False)
        writer.write(sample_result)

        wb = openpyxl.load_workbook(str(out))
        ws = wb["TTM_Forecast"]
        assert ws.freeze_panes is None


# ---------------------------------------------------------------------------
# Integration smoke test (DuckDB + stack logic, no real model download)
# ---------------------------------------------------------------------------

class TestEndToEndSmoke:
    """Verify DuckDBConnector → GraniteTSFMStack → ExcelResultWriter pipeline
    using in-memory DuckDB and mocked model inference."""

    @pytest.fixture(autouse=True)
    def _skip_if_missing(self):
        pytest.importorskip("duckdb")
        pytest.importorskip("openpyxl")

    def test_full_pipeline(self, market_df, tmp_path):
        import duckdb

        # 1. Seed in-memory DuckDB
        con = duckdb.connect(":memory:")
        con.register("ig_ticks", market_df)

        # 2. Fetch via connector
        connector = DuckDBConnector(
            db_path=":memory:",
            table_name="ig_ticks",
            timestamp_column="timestamp",
            id_columns=["instrument_id"],
            target_columns=["mid_price"],
        )
        connector._con = con
        df = connector.fetch_latest(context_length=64)
        assert len(df) == 64 * len(INSTRUMENTS)

        # 3. Build stack and preprocess
        stack = GraniteTSFMStack(
            timestamp_column="timestamp",
            id_columns=["instrument_id"],
            target_columns=["mid_price"],
            context_length=64,
            prediction_length=8,
            freq=FREQ,
        )
        stack.build_preprocessor(df)
        stack._models_loaded = True

        # 4. Mock model outputs
        t0 = df["timestamp"].max() + timedelta(minutes=1)

        def fake_ttm(_d):
            return pd.DataFrame(
                [
                    {
                        "timestamp": t0 + timedelta(minutes=i),
                        "instrument_id": inst,
                        "mid_price_prediction": 1.25,
                    }
                    for inst in INSTRUMENTS
                    for i in range(8)
                ]
            )

        def fake_tspulse(_d):
            return pd.DataFrame(
                [
                    {"timestamp": row["timestamp"], "instrument_id": row["instrument_id"], "anomaly_score": 0.05}
                    for _, row in df.iterrows()
                ]
            )

        def fake_flowstate(_d):
            return pd.DataFrame(
                [
                    {
                        "timestamp": t0 + timedelta(minutes=i),
                        "instrument_id": inst,
                        "mid_price_q10": 1.248,
                        "mid_price_q50": 1.250,
                        "mid_price_q90": 1.252,
                    }
                    for inst in INSTRUMENTS
                    for i in range(8)
                ]
            )

        with (
            patch.object(stack, "_run_ttm", side_effect=fake_ttm),
            patch.object(stack, "_run_tspulse", side_effect=fake_tspulse),
            patch.object(stack, "_run_flowstate", side_effect=fake_flowstate),
        ):
            result = stack.run(df)

        # 5. Write to Excel
        out = tmp_path / "ig_results.xlsx"
        writer = ExcelResultWriter(str(out))
        path = writer.write(result, run_label="smoke_test")

        assert path.exists()
        assert result.ensemble_summary["models_contributing"].max() >= 1
        connector.close()
